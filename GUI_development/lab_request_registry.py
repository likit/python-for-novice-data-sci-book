"""You will need to install the following libraries:
- PySimpleGUI
- openpyxl

"""


import PySimpleGUI as sg
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

sg.theme('BlueMono')

if not os.path.exists('demo.xlsx'):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['name', 'test', 'order_datetime', 'received_datetime'])
else:
    workbook = load_workbook(filename='demo.xlsx')
    worksheet = workbook.active


layout = [
    [sg.Text('Patient name:'), sg.Input(key='name', size=(40, 1))],
    [sg.Text('Test:')],
    [sg.Combo(values=['GLUC', 'CHOL', 'TG'], key='testcode', readonly=True)],
    [sg.Text('Ordered at:'), sg.Input(key='ordered_at'),
     sg.CalendarButton('Choose', target='ordered_at')],
    [sg.Button('Add'), sg.Button('View'), sg.Exit()]
]

window = sg.Window('Lab Request Registry', layout)


def view_data():
    rows = list(worksheet.values)
    headers = rows[0]
    data = rows[1:]
    layout = [
        [sg.Table(values=data, headings=headers)],
        [sg.Exit()],
    ]
    data_window = sg.Window('Data Window', layout)
    while True:
        event, _ = data_window.Read()
        if event in (sg.WIN_CLOSED, 'Exit'):
            break
    data_window.close()


def is_form_valid(values):
    """Complicated validation can be placed here."""
    if all([values[k] for k in ['name', 'testcode', 'ordered_at']]):
        return True
    return False


def clear_form(*fields):
    for field in fields:
        window[field].Update('')


while True:
    event, values = window.Read()

    if event in (sg.WIN_CLOSED, 'Exit'):
        break
    elif event == 'Add':
        if is_form_valid(values):
            worksheet.append([values['name'],
                              values['testcode'],
                              datetime.strptime(values['ordered_at'], '%Y-%m-%d %H:%M:%S'),
                              datetime.now()])
            clear_form('name', 'testcode', 'ordered_at')
        else:
            sg.PopupError('Form is not valid. You need to enter all information.')
    elif event == 'View':
        view_data()

window.close()
workbook.save(filename='demo.xlsx')
